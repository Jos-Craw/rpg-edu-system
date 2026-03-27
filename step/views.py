from django.shortcuts import render, get_object_or_404, redirect
from django.db.models import Count 
from .models import Group, Student, Lesson, Attendance, Performance
import datetime
from django.db.models import Avg, Count, Sum
from django.utils import timezone
import json
from .utils import get_xp_for_next_level, get_level_thresholds
from collections import Counter
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference, PieChart
from openpyxl.styles import Font
from collections import Counter

from .models import Group, Performance



def home(request):
    groups = Group.objects.all()
    all_lessons = Lesson.objects.all().order_by('-date')
    
    return render(request, 'home.html', {
        'groups': groups,            
        'all_lessons': all_lessons,   
    })



def group_detail(request, id):
    group = get_object_or_404(Group, id=id)
    students = group.student_set.all()
    lessons = group.lesson_set.all().order_by('date')
    
    line_series = []
    for student in students:
        total_xp = 0
        points = []
        for lesson in lessons:
            perf = Performance.objects.filter(student=student, lesson=lesson).first()
            if perf.classwork_score:
                total_xp += perf.classwork_score * 2

            if perf.homework_score:
                total_xp += perf.homework_score * 3
            points.append({'x': lesson.date.strftime('%d.%m'), 'y': total_xp})
        line_series.append({'name': student.name, 'data': points})

    ranks = [s.rank for s in students]
    rank_count = Counter(ranks)
    donut_labels = list(rank_count.keys())
    donut_values = list(rank_count.values())

    return render(request, 'group.html', {
        'group': group,
        'students': students,
        'lessons': lessons,
        'line_chart_data': json.dumps(line_series),
        'donut_labels': json.dumps(donut_labels),
        'donut_values': json.dumps(donut_values),
    })

def log_lesson(request, group_id):
    group = get_object_or_404(Group, id=group_id)
    students = group.student_set.all()
    
    if request.method == "POST":
        lesson = Lesson.objects.create(
            group=group,
            date=request.POST.get('date') or datetime.date.today(),
            topic=request.POST.get('topic', 'Без темы'),
            homework_description=request.POST.get('hw_desc', ''),
            topic_file=request.FILES.get('topic_file'),
            homework_file=request.FILES.get('homework_file')
        )
        
        attendances_to_create = []
        performances_to_create = []
        
        for student in students:
            is_present = request.POST.get(f'present_{student.id}') == 'on'
            hw_val = int(request.POST.get(f'hw_{student.id}', 0))
            cw_val = int(request.POST.get(f'cw_{student.id}', 0))
            
            attendances_to_create.append(
                Attendance(student=student, lesson=lesson, status=is_present)
            )
            performances_to_create.append(
                Performance(
                    student=student, 
                    lesson=lesson, 
                    homework_score=hw_val, 
                    classwork_score=cw_val
                )
            )
            
        Attendance.objects.bulk_create(attendances_to_create)
        Performance.objects.bulk_create(performances_to_create)
        for student in students:
            student.recalculate_stats()
            
        return redirect('group_detail', id=group.id)

    return render(request, 'log_lesson.html', {
        'group': group, 
        'students': students, 
        'today': datetime.date.today()
    })



from django.shortcuts import render, get_object_or_404
from django.db.models import Avg
from .models import Student, Performance, Lesson


def student_profile(request, student_id):
    student = get_object_or_404(Student, id=student_id)

    total_xp = student.xp
    level = student.level

    def get_level_thresholds(level):
        total = 0

        for lvl in range(level):
            total += get_xp_for_next_level(lvl)

        start = total
        end = total + get_xp_for_next_level(level)

        return start, end

    start_xp, end_xp = get_level_thresholds(student.level)

    current = student.xp - start_xp
    needed = end_xp - start_xp

    # защита
    if current < 0:
        current = 0
    if current > needed:
        current = needed

    progress_percent = int((current / needed) * 100) if needed > 0 else 0

    perf_query = Performance.objects.filter(student=student)

    avg_class = perf_query.aggregate(Avg('classwork_score'))['classwork_score__avg'] or 0
    avg_hw = perf_query.aggregate(Avg('homework_score'))['homework_score__avg'] or 0

    total_lessons = Lesson.objects.filter(group=student.group).count()
    attended_count = perf_query.count()
    attendance = (attended_count / total_lessons * 100) if total_lessons > 0 else 0

    stats_data = [
        round(avg_class * 10),
        round(avg_hw * 10),
        round(attendance),
        progress_percent,
        min(total_xp, 100)
    ]

    performances = perf_query.order_by('-lesson__date')[:10]

    current = student.xp - start_xp
    needed = end_xp - start_xp


    return render(request, 'student_profile.html', {
        'student': student,
        'current_xp': total_xp,
        'level_start': start_xp,
        'level_end': end_xp,
        'progress_percent': progress_percent,
        'performances': performances,
        'stats_data': stats_data,
        'current': current,
        'needed': needed,
        'progress_percent': progress_percent,
    })







def all_lessons_view(request):
    lessons = Lesson.objects.all().order_by("-date")
    return render(request, 'all_lessons.html', {'all_lessons': lessons})

def lesson_detail(request, lesson_id):
    lesson = get_object_or_404(Lesson, id=lesson_id)
    performances = Performance.objects.filter(lesson=lesson).select_related('student')
    
    for perf in performances:
        att = Attendance.objects.filter(lesson=lesson, student=perf.student).first()
        perf.is_present = att.status if att else False

    return render(request, 'lesson_detail.html', {
        'lesson': lesson,
        'performances': performances,
    })


def export_full_xlsx(request):
    wb = Workbook()
    wb.remove(wb.active)

    groups = Group.objects.all()

    for group in groups:
        ws = wb.create_sheet(title=group.name[:30])

        headers = ["Дата"]
        students = list(group.student_set.all())
        lessons = list(group.lesson_set.all().order_by('date'))

        for s in students:
            headers.append(s.name)

        ws.append(headers)

        for col in ws[1]:
            col.font = Font(bold=True)

        student_totals = {s.id: 0 for s in students}

        row_index = 2

        for lesson in lessons:
            row = [lesson.date.strftime('%d.%m')]

            for s in students:
                perf = Performance.objects.filter(student=s, lesson=lesson).first()

                xp = 0
                if perf:
                    xp = (perf.classwork_score or 0) * 2 + (perf.homework_score or 0) * 3

                student_totals[s.id] += xp
                row.append(student_totals[s.id])

            ws.append(row)
            row_index += 1

        chart = LineChart()
        chart.title = f"Прогресс XP — {group.name}"

        data = Reference(ws, min_col=2, max_col=len(students)+1, min_row=1, max_row=row_index-1)
        cats = Reference(ws, min_col=1, min_row=2, max_row=row_index-1)

        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)

        ws.add_chart(chart, f"A{row_index + 2}")

        rank_list = [s.rank for s in students]
        rank_count = Counter(rank_list)

        pie_start = row_index + 20

        ws[f"A{pie_start}"] = "Ранг"
        ws[f"B{pie_start}"] = "Кол-во"

        i = pie_start + 1
        for rank, count in rank_count.items():
            ws[f"A{i}"] = rank
            ws[f"B{i}"] = count
            i += 1

        pie = PieChart()
        labels = Reference(ws, min_col=1, min_row=pie_start+1, max_row=i-1)
        data = Reference(ws, min_col=2, min_row=pie_start, max_row=i-1)

        pie.add_data(data, titles_from_data=True)
        pie.set_categories(labels)

        ws.add_chart(pie, f"D{pie_start}")

        # --- 🏆 ТОП СТУДЕНТОВ ---
        top_start = pie_start + 15
        ws[f"A{top_start}"] = "ТОП студентов"
        ws[f"A{top_start}"].font = Font(bold=True)

        top_students = sorted(students, key=lambda x: x.xp, reverse=True)[:5]

        for idx, s in enumerate(top_students, start=1):
            ws[f"A{top_start + idx}"] = f"{idx}. {s.name}"
            ws[f"B{top_start + idx}"] = s.xp

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=analytics.xlsx'

    wb.save(response)
    return response